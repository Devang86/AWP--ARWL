# excel_generator.py  –  KKC Audit Work Program Excel Generator  v4.0
# Generates multi-sheet KKC-branded Excel workbook
# Ind AS updated | Detailed Revenue | NFRA Focus | No risk badges

import io
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from workprogram_data_v4 import WORKPROGRAMS, ASSERTION_MAP

KKC_GREEN  = "7CB542"; KKC_LGREEN = "D4EDAF"; KKC_GREY  = "808285"
KKC_LGREY  = "F4F4F4"; KKC_WHITE  = "FFFFFF"; KKC_DARK  = "1F2D3D"
KKC_BLUE   = "2C3E6B"; KKC_LBLUE  = "EBF0FF"; KKC_AMBER = "FFF8E1"
KKC_AMBER_FG = "5D4037"; KKC_RED  = "C0392B"

THIN = Side(border_style="thin",  color="CCCCCC")
MED  = Side(border_style="medium",color=KKC_GREY)

def fill(c):  return PatternFill("solid", fgColor=c)
def tb():     return Border(left=THIN,right=THIN,top=THIN,bottom=THIN)
def mb():     return Border(left=MED, right=MED, top=MED, bottom=MED)

def _merge(ws,r1,c1,r2,c2,value,bold=False,sz=10,bg=None,fg=KKC_WHITE,
           align="center",italic=False):
    ws.merge_cells(start_row=r1,start_column=c1,end_row=r2,end_column=c2)
    c = ws.cell(row=r1,column=c1,value=value)
    c.font = Font(name="Calibri",size=sz,bold=bold,color=fg,italic=italic)
    c.alignment = Alignment(horizontal=align,vertical="center",wrap_text=True)
    if bg: c.fill = fill(bg)
    c.border = mb()
    return c

def _assertion_labels(codes):
    from workprogram_data_v4 import ASSERTION_MAP
    return "  |  ".join(ASSERTION_MAP.get(c,c) for c in codes)

def _rows(text,chars=85):
    if not text: return 2
    return max(sum(max(1,(len(l)//chars)+1) for l in text.split("\n")),2)

COLUMNS = [
    ("Step\nNo.",                         7),
    ("Category",                         22),
    ("Audit Risk\n(Ind AS Basis)",        40),
    ("Audit Procedure\n(Actionable Steps)", 72),
    ("Assertions\nAddressed",             28),
    ("SA Reference\n(Para No.)",          36),
    ("Ind AS / Regulatory\nReference",    30),
    ("KKC Manual\nRef.",                  16),
    ("Responsible",                       14),
    ("Status",                            12),
    ("WP Reference",                      16),
    ("Reviewer Comments",                 28),
]


def create_cover(wb, company, period, areas, materiality, ep, em):
    ws = wb.create_sheet("Cover Page", 0)
    ws.sheet_view.showGridLines = False
    for i in range(1,13): ws.column_dimensions[get_column_letter(i)].width = 15
    ws.row_dimensions[2].height = 55
    _merge(ws,2,1,2,12,"KKC & ASSOCIATES LLP\nCHARTERED ACCOUNTANTS",
           bold=True,sz=22,bg=KKC_GREEN)
    ws.row_dimensions[4].height = 30
    _merge(ws,4,1,4,12,"STATUTORY AUDIT — DETAILED AUDIT WORK PROGRAM  |  Ind AS Updated v4.0",
           bold=True,sz=13,bg=KKC_BLUE)
    ws.row_dimensions[5].height = 20
    _merge(ws,5,1,5,12,"Mumbai  ·  Pune  ·  Bengaluru  ·  Ahmedabad",
           sz=10,bg=KKC_GREY,italic=True)

    total_steps = sum(len(WORKPROGRAMS.get(a,{}).get("steps",[])) for a in areas)
    details = [
        ("Client Name",           company or "—"),
        ("Period of Audit",       period or "—"),
        ("Date Generated",        datetime.today().strftime("%d %B %Y")),
        ("Engagement Partner",    ep or "To be assigned"),
        ("Engagement Manager",    em or "To be assigned"),
        ("Overall Materiality",   materiality or "To be determined"),
        ("Performance Materiality","75% of Overall Materiality"),
        ("Audit Areas",           f"{len(areas)} areas selected"),
        ("Total Procedures",      f"{total_steps} detailed steps"),
        ("Standards",             "ICAI SAs | All Ind AS (MCA) | KKC Audit Manual 2026"),
        ("Revenue Coverage",      "MF Distribution (RTA reconciliation CAMS+KFintech) + Structured Product Gains (Ind AS 109)"),
        ("NFRA Compliance",       "Dedicated NFRA Focus Area – based on published inspection reports 2021-2024"),
    ]
    row = 7
    for label, val in details:
        ws.row_dimensions[row].height = 20
        _merge(ws,row,1,row,4,label,bold=True,bg=KKC_LGREY,fg=KKC_DARK,align="left")
        _merge(ws,row,5,row,12,val,bg=KKC_WHITE,fg=KKC_DARK,align="left")
        row += 1

    row += 1
    _merge(ws,row,1,row,12,"AUDIT AREAS INCLUDED IN THIS WORK PROGRAM",
           bold=True,sz=11,bg=KKC_GREEN); row += 1
    for idx, area in enumerate(areas, 1):
        n = len(WORKPROGRAMS.get(area,{}).get("steps",[]))
        bg_c = KKC_LGREEN if idx%2==0 else KKC_WHITE
        ws.row_dimensions[row].height = 18
        _merge(ws,row,1,row,1,str(idx),bg=bg_c,fg=KKC_DARK)
        _merge(ws,row,2,row,10,area,bg=bg_c,fg=KKC_DARK,align="left")
        _merge(ws,row,11,row,12,f"{n} steps",bg=bg_c,fg=KKC_GREY,align="center")
        row += 1

    row += 2
    ws.row_dimensions[row].height = 55
    _merge(ws,row,1,row,12,
           "DISCLAIMER: This Audit Work Program is prepared based on ICAI SAs, Indian Accounting Standards "
           "(Ind AS) as notified by MCA (sourced from All_IND_AS_merged.pdf), KKC Audit Manual 2026, and "
           "publicly available information on the auditee. Procedures must be adapted to actual risk assessment, "
           "materiality, and entity-specific circumstances by the engagement team. "
           "No risk classification badges are applied – all procedures require execution.",
           italic=True,sz=8,bg=KKC_AMBER,fg=KKC_AMBER_FG,align="left")


def create_summary(wb, areas, company, period):
    ws = wb.create_sheet("Summary", 1)
    ws.sheet_view.showGridLines = False
    for i,w in enumerate([6,32,8,8,34,14,14,16],1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[2].height = 40
    _merge(ws,2,1,2,8,
           f"SUMMARY  |  {company}  |  {period}",bold=True,sz=13,bg=KKC_GREEN)
    ws.row_dimensions[4].height = 28
    for ci,h in enumerate(["#","Audit Area","Steps","Key Ind AS","Key SAs","Status","Prepared By","Date"],1):
        c = ws.cell(row=4,column=ci,value=h)
        c.font = Font(name="Calibri",size=9,bold=True,color=KKC_WHITE)
        c.fill = fill(KKC_DARK); c.border = mb()
        c.alignment = Alignment(horizontal="center",vertical="center",wrap_text=True)
    row = 5
    for idx,area in enumerate(areas,1):
        steps = WORKPROGRAMS.get(area,{}).get("steps",[])
        ind_as = set(); sa_nums = set()
        for s in steps:
            for p in s.get("sa_ref","").split("|"):
                p = p.strip()
                if p.startswith("Ind AS"): ind_as.add(p.split("Para")[0].strip()[:14])
                elif p.startswith("SA"):    sa_nums.add(" ".join(p.split()[:2])[:9])
        bg_c = KKC_LGREEN if idx%2==1 else KKC_WHITE
        ws.row_dimensions[row].height = 22
        for ci,val in enumerate([idx,area,len(steps),
                                  ", ".join(sorted(ind_as)[:3]),
                                  ", ".join(sorted(sa_nums)[:4]),
                                  "Pending","",""],1):
            c = ws.cell(row=row,column=ci,value=val)
            c.font = Font(name="Calibri",size=9,color=KKC_DARK)
            c.fill = fill(bg_c)
            c.alignment = Alignment(horizontal="center" if ci in(1,3,6) else "left",
                                    vertical="center",wrap_text=True)
            c.border = tb()
        row += 1
    ws.freeze_panes = "A5"


def create_indas_sheet(wb):
    ws = wb.create_sheet("Ind AS Modifications", 2)
    ws.sheet_view.showGridLines = False
    for i,w in enumerate([5,18,42,52,22],1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[2].height = 34
    _merge(ws,2,1,2,5,
           "IND AS MODIFICATIONS TO AUDIT WORKPROGRAMS  |  Sourced from All_IND_AS_merged.pdf (MCA, 1591 pages)",
           bold=True,sz=12,bg=KKC_GREEN)
    ws.row_dimensions[3].height = 18
    _merge(ws,3,1,3,5,
           "Key Ind AS requirements reviewed and reflected in the audit workprograms. Engagement team should read relevant Ind AS before finalising procedures.",
           sz=9,bg=KKC_BLUE,fg=KKC_WHITE,italic=True)
    ws.row_dimensions[4].height = 28
    for ci,h in enumerate(["#","Ind AS","Key Requirement (Para Ref.)","Workprogram Modification","SA Linkage"],1):
        c = ws.cell(row=4,column=ci,value=h)
        c.font = Font(name="Calibri",size=9,bold=True,color=KKC_WHITE)
        c.fill = fill(KKC_DARK); c.border = mb()
        c.alignment = Alignment(horizontal="center",vertical="center",wrap_text=True)

    mods = [
        ("Ind AS 115\nPara 9,22,35-38,47,56-58",
         "Revenue from Contracts with Customers. Trail commission = variable consideration (AUM-linked, Para 50). "
         "Apply constraint (Para 56-58): recognise only if highly probable no significant reversal. "
         "Trail = over-time PO (Para 35 – customer simultaneously receives/consumes benefit). "
         "Upfront = point-in-time PO (Para 38 – investment execution). "
         "Accrued income = contract asset (Para 116).",
         "Revenue area expanded to 15 steps (9 MF Distribution + 6 Structured Products). "
         "Step 1: Five-step model applied explicitly. Step 5: AUM verification vs AMFI published. "
         "Step 7: Contract asset (accrued income) reconciled. Upfront AMFI ban compliance verified. "
         "Variable consideration constraint documented.",
         "SA 240 Para 26 | SA 500 | SA 520"),

        ("Ind AS 109\nPara 4.1.1-4.1.4, 3.2.3, 3.2.6, 3.2.12, B5.4.2",
         "Financial Instruments. Classification: Amortised Cost (SPPI + HTC), FVOCI (SPPI + HTC&S), "
         "FVTPL (residual). SPPI test (Para 4.1.3): solely P&I – MLDs fail (market-linked return). "
         "Derecognition (Para 3.2.3): rights expire or qualifying transfer (risks and rewards per 3.2.6). "
         "Gain on derecognition (Para 3.2.12): consideration received minus carrying amount. "
         "Dirty price separation (B5.4.2): accrued interest distinct from capital gain.",
         "Structured product steps added: SPPI test documented per instrument (Step 10). "
         "Gain computation includes dirty/clean price separation (Step 12B). "
         "MTM anti-double-count check added (Step 12C). Settlement before derecognition verified (Step 13C). "
         "Investments: SPPI docs required. ECL (Chapter 5.5) assessment added.",
         "SA 500 | SA 540 | SA 620"),

        ("Ind AS 37\nPara 14-16, 27-30",
         "Provisions: three cumulative conditions – (a) present obligation from past event, "
         "(b) probable outflow of economic resources, (c) reliable estimate of amount. "
         "Contingent liability: possible obligation or probable but no reliable estimate – DISCLOSE ONLY. "
         "Remote: no recognition, no disclosure. Interaction: financial guarantees → Ind AS 109 (not Ind AS 37).",
         "Provisions workprogram: each provision tested against Para 14 three conditions. "
         "Contingent liabilities: Para 27-30 decision tree applied. "
         "Financial guarantee contracts routed to Ind AS 109.",
         "SA 501 Para 9 | SA 540"),

        ("Ind AS 19\nPara 1, 83, 120-152",
         "Employee Benefits. Actuarial gains/losses → OCI (remeasurement). "
         "Current service cost + net interest cost (DBO × discount rate) → P&L. "
         "Discount rate = yield on government bonds at balance sheet date (Para 83). "
         "ESOP/LTIP → Ind AS 102 (separate standard). "
         "Short-term benefits (< 12 months): undiscounted accrual.",
         "Actuarial valuation: discount rate verified vs G-Sec yield (March 2026). "
         "OCI vs P&L presentation verified. Ind AS 102 ESOP kept separate. "
         "Auditor's own DBO estimate developed per SA 540 Para 13 (NFRA requirement).",
         "SA 540 | SA 620"),

        ("Ind AS 116\nPara 18-20, 26-29, 5-6",
         "Leases. Lease term: non-cancellable period + renewals if 'reasonably certain' to exercise (Para 18-19). "
         "IBR = rate lessee would pay to borrow funds of similar amount, term, security (Para 26). "
         "Reassess if significant event changes likelihood (Para 20). "
         "Short-term (≤12 months original term) and low-value: optional exemption (Para 5-6). "
         "SaaS: not within Ind AS 116 scope (no right to control underlying asset) → expense.",
         "Lease term documentation expanded: 'reasonably certain' analysis for ARWL office leases. "
         "IBR benchmarked to ARWL secured borrowing rates. "
         "Reassessment events identified. SaaS confirmed as expense (not lease).",
         "SA 500 | SA 540"),

        ("Ind AS 12\nPara 15-18, 24, 47",
         "Income Taxes. Deferred tax on ALL temporary differences (balance sheet approach). "
         "DTA recognised if probable future taxable profit (Para 24). "
         "Deferred tax on OCI items → OCI. On equity items → equity. "
         "Tax rate = enacted/substantively enacted rate at balance sheet date (Para 47). "
         "ESOP: DTA when exercise proceeds less than P&L charge (deductible on exercise).",
         "Taxation workprogram: all temporary differences identified (Sch II depreciation, "
         "ESOP, DBO actuarial, ECL provisions, FVTPL MTM). "
         "DTA recoverability assessed given ARWL PAT of Rs.301 Cr (probable future profits). "
         "OCI deferred tax verified.",
         "SA 540 | SA 500"),

        ("Ind AS 24\nPara 9, 17-21",
         "Related Party Disclosures. Definition (Para 9): KMP, close family, entities controlled by KMP, "
         "subsidiaries, associates, joint ventures, post-employment benefit plans. "
         "KMP compensation (Para 17): disclose in total by category (ST benefits, LT benefits, ESOP, termination, post-employment). "
         "All RPTs disclosed even if at arm's length – no exemption for 'normal terms'.",
         "RPT workprogram: Ind AS 24 Para 9 applied for identification. "
         "KMP compensation note cross-checked per Para 17. "
         "SEBI LODR Reg 23 AC approval verified. "
         "Structured product transactions with group entities tested for arm's length.",
         "SA 550 | SA 315"),

        ("Ind AS 36\nPara 9-12, 59, 66",
         "Impairment. Annual test: goodwill and indefinite-life intangibles (Para 10). "
         "Indicator-based: other assets (Para 9). "
         "Recoverable amount = max(FVLCTS, VIU). "
         "VIU: pre-tax DCF using pre-tax discount rate (Para 55-56). "
         "CGU identification required (Para 66).",
         "Investments workprogram: Ind AS 36 indicators assessed for subsidiaries/associates. "
         "If indicators present: DCF audit of management's VIU computation. "
         "Fixed assets: impairment indicators reviewed annually.",
         "SA 540 | SA 620"),
    ]

    row = 5
    for idx,(ind_as,req,impact,sa) in enumerate(mods,1):
        h = min(max(_rows(req,55)*13, _rows(impact,62)*13, 40), 160)
        ws.row_dimensions[row].height = h
        bg_c = KKC_LGREEN if idx%2==1 else KKC_WHITE
        for ci,val in enumerate([idx,ind_as,req,impact,sa],1):
            c = ws.cell(row=row,column=ci,value=str(val))
            c.font = Font(name="Calibri",size=9,color=KKC_DARK,bold=(ci==2))
            c.fill = fill(bg_c)
            c.alignment = Alignment(horizontal="left" if ci>1 else "center",
                                    vertical="top",wrap_text=True)
            c.border = tb()
        row += 1
    ws.freeze_panes = "A5"


def create_area_sheet(wb, area_name, area_data, company, period):
    safe = area_name.replace("/","-").replace(":","").replace("–","-")[:31]
    ws = wb.create_sheet(safe)
    ws.sheet_view.showGridLines = False
    for i,(_,w) in enumerate(COLUMNS,1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.row_dimensions[1].height = 8
    ws.row_dimensions[2].height = 28
    _merge(ws,2,1,2,len(COLUMNS),
           "KKC & ASSOCIATES LLP  |  STATUTORY AUDIT WORK PROGRAM  |  Ind AS Updated",
           bold=True,sz=11,bg=KKC_GREEN)
    ws.row_dimensions[3].height = 20
    _merge(ws,3,1,3,len(COLUMNS),
           f"Area: {area_name}   |   Client: {company}   |   Period: {period}",
           bold=True,sz=10,bg=KKC_BLUE)

    ws.row_dimensions[5].height = 14
    _merge(ws,5,1,5,3,"AUDIT OBJECTIVE",bold=True,bg=KKC_GREEN,fg=KKC_WHITE,align="left")
    _merge(ws,5,4,5,len(COLUMNS),"",bg=KKC_GREEN)
    ws.row_dimensions[6].height = max(_rows(area_data["objective"],110)*14, 40)
    _merge(ws,6,1,6,len(COLUMNS),area_data["objective"],bg=KKC_LBLUE,fg=KKC_DARK,align="left",sz=9)

    ws.row_dimensions[7].height = 14
    _merge(ws,7,1,7,3,"RISK OVERVIEW (Ind AS basis)",bold=True,bg=KKC_GREY,fg=KKC_WHITE,align="left")
    _merge(ws,7,4,7,len(COLUMNS),"",bg=KKC_GREY)
    ws.row_dimensions[8].height = max(_rows(area_data["risk_overview"],110)*14, 50)
    _merge(ws,8,1,8,len(COLUMNS),area_data["risk_overview"],
           bg=KKC_AMBER,fg=KKC_AMBER_FG,align="left",sz=9,italic=True)

    ws.row_dimensions[9].height = 8
    ws.row_dimensions[10].height = 42
    for ci,(col_name,_) in enumerate(COLUMNS,1):
        c = ws.cell(row=10,column=ci,value=col_name)
        c.font = Font(name="Calibri",size=9,bold=True,color=KKC_WHITE)
        c.fill = fill(KKC_BLUE)
        c.alignment = Alignment(horizontal="center",vertical="center",wrap_text=True)
        c.border = mb()

    data_start = 11
    for si,step in enumerate(area_data["steps"]):
        row = data_start + si
        alt = (si%2==1)
        row_bg = KKC_LGREY if alt else KKC_WHITE
        cat_bg = KKC_LGREEN if alt else "E8F5E9"
        sa_bg  = KKC_LBLUE  if alt else "EBF0FF"

        ws.row_dimensions[row].height = min(max(_rows(step["procedure"],82)*13,55), 180)

        assertions_text = _assertion_labels(step["assertions"])
        full_ref = step.get("sa_ref","")
        sa_parts    = [p.strip() for p in full_ref.split("|")
                       if p.strip().startswith("SA") or p.strip().startswith("SQC") or p.strip().startswith("CARO")]
        ind_as_parts= [p.strip() for p in full_ref.split("|")
                       if p.strip().startswith("Ind AS") or p.strip().startswith("SEBI") or
                          p.strip().startswith("Companies") or p.strip().startswith("Rule")]

        row_data = [
            (step["no"],          row_bg,      KKC_DARK,     "center"),
            (step["category"],    cat_bg,      KKC_BLUE,     "left"),
            (step["risk"],        KKC_AMBER,   KKC_AMBER_FG, "left"),
            (step["procedure"],   row_bg,      KKC_DARK,     "left"),
            (assertions_text,     row_bg,      KKC_BLUE,     "left"),
            (" | ".join(sa_parts),    sa_bg,   KKC_DARK,     "left"),
            (" | ".join(ind_as_parts),sa_bg,   "1A5276",     "left"),
            (step["kkc_ref"],     row_bg,      KKC_GREY,     "left"),
            ("Article / AM / EP", row_bg,      KKC_GREY,     "center"),
            ("Pending",           row_bg,      KKC_GREY,     "center"),
            ("",                  row_bg,      KKC_DARK,     "center"),
            ("",                  row_bg,      KKC_DARK,     "left"),
        ]
        for ci,(val,bg_c,fg_c,aln) in enumerate(row_data,1):
            c = ws.cell(row=row,column=ci,value=str(val) if val else "")
            c.font = Font(name="Calibri",size=9,color=fg_c,bold=(ci==2))
            c.alignment = Alignment(horizontal=aln,vertical="top",wrap_text=True)
            c.fill = fill(bg_c); c.border = tb()

    sign_row = data_start + len(area_data["steps"]) + 1
    ws.row_dimensions[sign_row].height = 22
    sign_items = ["Prepared by:","","Date:","",
                  "Reviewed by (AM):","","Date:","",
                  "Reviewed by (EP):","","Date:",""]
    for ci,lbl in enumerate(sign_items[:len(COLUMNS)],1):
        c = ws.cell(row=sign_row,column=ci,value=lbl)
        c.font = Font(name="Calibri",size=8,italic=True,color=KKC_GREY)
        c.fill = fill(KKC_LGREY); c.border = tb()
        c.alignment = Alignment(horizontal="left",vertical="center")

    ws.freeze_panes = "A11"


def generate_excel(company, period, areas, materiality="", ep="", em=""):
    wb = openpyxl.Workbook()
    if "Sheet" in wb.sheetnames: del wb["Sheet"]

    create_cover(wb, company, period, areas, materiality, ep, em)
    create_summary(wb, areas, company, period)
    create_indas_sheet(wb)

    for area in areas:
        if area in WORKPROGRAMS:
            create_area_sheet(wb, area, WORKPROGRAMS[area], company, period)

    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    return buf.read()
